from pathlib import Path
import io
import re
import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook

st.set_page_config(
    page_title="Academic Product Margin Intelligence",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

APP_DIR = Path(__file__).resolve().parent
DEFAULT_FILE = APP_DIR / "data" / "margin_model.xlsx"

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Space+Grotesk:wght@500;600;700&display=swap');
:root {font-family:'Manrope',sans-serif;color:#172554;background:#eef4ff;}
html, body, [class*="css"] {font-family:'Manrope',sans-serif;letter-spacing:.005em;}
.stApp {background:radial-gradient(circle at 5% 0%,#dbeafe 0,#f8fafc 34%,#eef2ff 100%);color:#172554;}
.block-container {padding-top:1.4rem;padding-bottom:2.5rem;max-width:1500px;}
[data-testid="stSidebar"] {background:linear-gradient(160deg,#07122f 0%,#10265b 55%,#1d4ed8 100%);box-shadow:12px 0 35px rgba(15,23,42,.18);}
[data-testid="stSidebar"] * {color:#f8fafc;}
[data-testid="stSidebar"] .stMultiSelect span,[data-testid="stSidebar"] .stSelectbox span {color:#111827;}
[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] > div {background:#fff;border-color:#bfdbfe;}
[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] div {color:#0f172a !important;}
[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] svg {fill:#334155;}
[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] * {color:#0f172a !important;-webkit-text-fill-color:#0f172a !important;}
[data-testid="stSidebar"] [data-testid="stSelectbox"] input {color:#0f172a !important;-webkit-text-fill-color:#0f172a !important;caret-color:#0f172a;}
[data-baseweb="popover"] [role="option"] {color:#111827 !important;background:#fff !important;}
[data-baseweb="popover"] [role="option"] * {color:#111827 !important;-webkit-text-fill-color:#111827 !important;}
[data-testid="stSidebar"] .stMarkdown h2,[data-testid="stSidebar"] .stMarkdown h3 {font-family:'Space Grotesk',sans-serif;letter-spacing:-.02em;}
.hero {background:linear-gradient(120deg,#081533 0%,#123b8f 58%,#2563eb 100%);border:1px solid rgba(147,197,253,.35);border-radius:24px;padding:30px 34px;color:white;box-shadow:0 22px 0 rgba(15,23,42,.08),0 28px 55px rgba(30,64,175,.28);margin-bottom:22px;position:relative;overflow:hidden;transform:translateY(-2px);}
.hero:before {content:"";position:absolute;width:360px;height:360px;border:1px solid rgba(255,255,255,.18);border-radius:50%;right:-90px;top:-180px;box-shadow:0 0 0 28px rgba(255,255,255,.04),0 0 0 56px rgba(255,255,255,.03);}
.hero:after {content:"";position:absolute;width:180px;height:180px;border-radius:50%;right:80px;bottom:-135px;background:rgba(125,211,252,.22);filter:blur(2px);}
.hero h1 {font-family:'Space Grotesk',sans-serif;font-size:clamp(25px,3vw,36px);line-height:1.15;margin:0 0 9px;font-weight:700;letter-spacing:-.035em;position:relative;z-index:1;}
.hero p {margin:0;max-width:930px;color:#dbeafe;font-size:14px;line-height:1.65;position:relative;z-index:1;}
.eyebrow {font-size:11px;font-weight:800;letter-spacing:.16em;text-transform:uppercase;color:#7dd3fc;margin-bottom:10px;position:relative;z-index:1;}
.kpi {background:linear-gradient(145deg,rgba(255,255,255,.98),rgba(239,246,255,.94));border:1px solid rgba(148,163,184,.28);border-radius:18px;padding:18px 19px;box-shadow:0 8px 0 rgba(148,163,184,.15),0 14px 28px rgba(15,23,42,.1);min-height:126px;transition:transform .2s ease,box-shadow .2s ease;}
.kpi:hover {transform:translateY(-4px);box-shadow:0 12px 0 rgba(148,163,184,.14),0 22px 34px rgba(30,64,175,.15);}
.kpi-label {font-size:12px;color:#475569;font-weight:800;margin-bottom:8px;line-height:1.35;}.kpi-label.landing{color:#2563eb}.kpi-label.cost{color:#d97706}.kpi-value{font-family:'Space Grotesk',sans-serif;font-size:clamp(20px,2vw,28px);font-weight:700;color:#0f172a;letter-spacing:-.035em;white-space:nowrap}.kpi-foot{font-size:11px;color:#64748b;margin-top:9px;line-height:1.4}.good{color:#059669;font-weight:800}.warn{color:#d97706;font-weight:800}.bad{color:#dc2626;font-weight:800}
.section-title {font-family:'Space Grotesk',sans-serif;font-size:19px;font-weight:700;color:#172554;margin-top:6px;margin-bottom:3px;letter-spacing:-.02em;}.section-sub{font-size:12px;color:#64748b;margin-bottom:12px;line-height:1.5}
.info-card {background:linear-gradient(145deg,#fff,#eff6ff);border:1px solid rgba(148,163,184,.3);border-radius:17px;padding:16px 18px;box-shadow:0 7px 0 rgba(148,163,184,.13),0 13px 25px rgba(15,23,42,.08);min-height:108px;}
div[data-testid="stMetric"] {background:linear-gradient(145deg,#fff,#eff6ff);border:1px solid #dbeafe;padding:13px 16px;border-radius:15px;box-shadow:0 6px 0 rgba(148,163,184,.12),0 10px 20px rgba(15,23,42,.07);}
.stTabs [data-baseweb="tab-list"] {gap:7px;background:rgba(226,232,240,.82);padding:7px;border-radius:15px;box-shadow:inset 0 2px 5px rgba(15,23,42,.08);overflow-x:auto;}
.stTabs [data-baseweb="tab"] {height:42px;border-radius:10px;padding:0 16px;white-space:nowrap;font-weight:700;color:#475569;}
.stTabs [aria-selected="true"] {background:white;box-shadow:0 4px 0 rgba(148,163,184,.16),0 5px 12px rgba(15,23,42,.1);color:#1d4ed8;}
[data-testid="stDataFrame"] {border:1px solid #dbeafe;border-radius:15px;overflow:hidden;box-shadow:0 7px 0 rgba(148,163,184,.11),0 12px 22px rgba(15,23,42,.06);}
.stButton>button,.stDownloadButton>button {border-radius:10px;font-weight:700;border:1px solid #bfdbfe;box-shadow:0 4px 0 #bfdbfe;transition:transform .15s ease;}
.stButton>button:hover,.stDownloadButton>button:hover {transform:translateY(-2px);box-shadow:0 6px 0 #93c5fd;}
@media (max-width:760px) {.block-container{padding:1rem .8rem 2rem}.hero{padding:24px 22px}.kpi{min-height:112px}.kpi-value{font-size:20px}.stTabs [data-baseweb="tab"]{padding:0 11px;font-size:12px}}
/* Netflix-inspired cinematic palette */
:root {color:#f5f5f1;background:#141414;}
.stApp {background:radial-gradient(circle at 50% -15%,#3b1015 0,#181818 34%,#0f0f0f 100%);color:#f5f5f1;}
[data-testid="stSidebar"] {background:linear-gradient(180deg,#0b0b0b 0%,#171717 55%,#260b0e 100%);box-shadow:12px 0 35px rgba(0,0,0,.35);border-right:1px solid #292929;}
.hero {background:linear-gradient(112deg,#090909 0%,#1d1d1d 48%,#7f1018 100%);border:1px solid #4a171b;border-radius:5px;box-shadow:0 25px 60px rgba(0,0,0,.45);}
.hero:before {border-color:rgba(229,9,20,.35);box-shadow:0 0 0 35px rgba(229,9,20,.08),0 0 0 70px rgba(229,9,20,.04);}
.hero:after {background:rgba(229,9,20,.28);}
.hero p {color:#d2d2d2;}
.eyebrow {color:#e50914;}
.kpi {background:linear-gradient(145deg,#252525,#171717);border-color:#383838;box-shadow:0 12px 24px rgba(0,0,0,.28);}
.kpi:hover {box-shadow:0 16px 30px rgba(0,0,0,.42),0 0 0 1px #e50914;}
.kpi-label {color:#b3b3b3;}.kpi-label.landing{color:#f0444c}.kpi-label.cost{color:#f5a623}.kpi-value{color:#fff}.kpi-foot{color:#a3a3a3;}
.good{color:#46d369}.warn{color:#f5a623}.bad{color:#e50914}
.section-title {color:#fff}.section-sub {color:#a3a3a3}
.info-card,div[data-testid="stMetric"] {background:linear-gradient(145deg,#252525,#171717);border-color:#383838;box-shadow:0 12px 24px rgba(0,0,0,.28);color:#f5f5f1;}
.stTabs [data-baseweb="tab-list"] {background:#222;box-shadow:inset 0 2px 5px rgba(0,0,0,.4);}
.stTabs [data-baseweb="tab"] {color:#b3b3b3}.stTabs [aria-selected="true"] {background:#e50914;color:#fff;box-shadow:0 5px 14px rgba(229,9,20,.3);}
[data-testid="stDataFrame"] {border-color:#383838;box-shadow:0 12px 24px rgba(0,0,0,.24);}
.stButton>button,.stDownloadButton>button {border-color:#e50914;background:#e50914;color:#fff;box-shadow:0 4px 0 #8b050d;}
.stButton>button:hover,.stDownloadButton>button:hover {background:#f6121d;color:#fff;box-shadow:0 6px 0 #8b050d;}
.stPlotlyChart {background:#1b1b1b;border-radius:5px;padding:4px;}
</style>
""", unsafe_allow_html=True)


def clean_numeric(s):
    return pd.to_numeric(s, errors="coerce").fillna(0.0)


def extract_ay(value):
    if pd.isna(value):
        return None
    text = str(value)
    m = re.search(r"(20\d{2})\s*[-/]\s*(20\d{2})", text)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r"(?:AY\s*)?(\d{2})\s*[-/]\s*(\d{2})", text, flags=re.I)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return f"20{a:02d}-20{b:02d}"
    return None


def grade_num(value):
    if pd.isna(value): return 999
    m = re.search(r"(\d+)", str(value))
    return int(m.group(1)) if m else 999


def read_assumptions(source):
    if hasattr(source, "seek"): source.seek(0)
    wb = load_workbook(source, data_only=False, read_only=True)
    ws = wb["Assumptions"]
    target = ws["B3"].value or 0.30
    basis = ws["B13"].value or "Landing Price Excl GST"
    rates = {}
    for r in range(6, 11):
        p = ws.cell(r,1).value
        if p:
            rates[str(p)] = {
                "Transportation": float(ws.cell(r,2).value or 0),
                "Warehouse": float(ws.cell(r,3).value or 0),
                "Salary": float(ws.cell(r,4).value or 0),
                "Additional Services": float(ws.cell(r,5).value or 0),
            }
    wb.close()
    return float(target), str(basis), rates


def read_all_sheets(source):
    if hasattr(source, "seek"):
        source.seek(0)
    return pd.read_excel(source, sheet_name=None, engine="openpyxl")


def load_margin(source):
    if hasattr(source, "seek"): source.seek(0)
    df = pd.read_excel(source, sheet_name="Margin Analysis", engine="openpyxl")
    # Normalize core raw fields and independently recompute Excel formula outputs.
    for c in ["CP", "GST %", "Base Price with GST", "Base %", "MRP Bundle", "MRP IND"]:
        if c in df: df[c] = pd.to_numeric(df[c], errors="coerce")

    fixed = df.get("Value Type", pd.Series(index=df.index, dtype=object)).astype(str).str.strip().eq("Fixed Price ( Base Price with GST)")
    calc_mrp = np.where(
        fixed,
        df.get("Base Price with GST", 0),
        clean_numeric(df.get("Base %", 0)) * clean_numeric(df.get("MRP Bundle", 0)),
    )
    existing_mrp = pd.to_numeric(df.get("MRP IND", np.nan), errors="coerce")
    df["Calculated Landing Incl GST"] = existing_mrp.where(existing_mrp.notna(), pd.Series(calc_mrp, index=df.index)).fillna(0)
    df["GST %"] = clean_numeric(df.get("GST %", 0))
    df["Landing Price Excl GST (₹)"] = np.where(
        (1 + df["GST %"]) != 0,
        df["Calculated Landing Incl GST"] / (1 + df["GST %"]),
        df["Calculated Landing Incl GST"],
    )
    df["Cost Price (₹)"] = clean_numeric(df.get("CP", 0))
    df["Gross Margin (₹)"] = df["Landing Price Excl GST (₹)"] - df["Cost Price (₹)"]
    df["Gross Margin %"] = np.where(df["Landing Price Excl GST (₹)"] != 0, df["Gross Margin (₹)"] / df["Landing Price Excl GST (₹)"], 0)

    ed = df.get("Edition", pd.Series(index=df.index, dtype=object)).map(extract_ay)
    pn = df.get("Product Name", pd.Series(index=df.index, dtype=object)).map(extract_ay)
    df["Academic Year"] = ed.fillna(pn).fillna("Unknown")
    df["Product Type"] = df.get("Product Type", "Unknown").fillna("Unknown").astype(str).str.strip()
    df["Grades"] = df.get("Grades", "Unknown").fillna("Unknown").astype(str).str.strip()
    df["Subjects"] = df.get("Subjects", "Unknown").fillna("Unknown").astype(str).str.strip()
    df["Item Type"] = df.get("Item Type", "Unknown").fillna("Unknown").astype(str).str.strip()
    return df


def apply_model(df, rates, basis, target):
    out = df.copy()
    base = out["Cost Price (₹)"] if basis == "Cost Price" else out["Landing Price Excl GST (₹)"]
    for key, col in [("Transportation","Transportation Cost (₹)"),("Warehouse","Warehouse & Facility Cost (₹)"),("Salary","Employee Salary Cost (₹)"),("Additional Services","Additional Services Cost (₹)")]:
        pct = out["Product Type"].map(lambda x: rates.get(x, {}).get(key, 0.0)).fillna(0)
        out[col] = base * pct
    cost_cols = ["Transportation Cost (₹)","Warehouse & Facility Cost (₹)","Employee Salary Cost (₹)","Additional Services Cost (₹)"]
    out["Total Operating Cost (₹)"] = out[cost_cols].sum(axis=1)
    out["Net Margin (₹)"] = out["Gross Margin (₹)"] - out["Total Operating Cost (₹)"]
    out["Net Margin %"] = np.where(out["Landing Price Excl GST (₹)"] != 0, out["Net Margin (₹)"] / out["Landing Price Excl GST (₹)"], 0)
    out["Net Margin Status"] = np.where(out["Net Margin %"] >= target, "On / Above Target", "Below Target")
    return out


def money(v):
    if abs(v) >= 1e7: return f"₹{v/1e7:,.2f} Cr"
    if abs(v) >= 1e5: return f"₹{v/1e5:,.2f} L"
    return f"₹{v:,.0f}"


def pct(v): return f"{v*100:.1f}%"


def kpi(label, value, foot, state="", label_class=""):
    cls = {"good":"good","warn":"warn","bad":"bad"}.get(state,"")
    st.markdown(f'<div class="kpi"><div class="kpi-label {label_class}">{label}</div><div class="kpi-value">{value}</div><div class="kpi-foot {cls}">{foot}</div></div>', unsafe_allow_html=True)

@st.cache_data(show_spinner=False)
def load_default_bytes(path_str):
    return Path(path_str).read_bytes()

st.sidebar.markdown("## Margin Intelligence")
st.sidebar.caption("Academic Product Profitability • Executive Analytics")
upload = st.sidebar.file_uploader("Use another Excel model", type=["xlsx"], help="Expected sheets: Margin Analysis and Assumptions")
source_bytes = upload.getvalue() if upload else load_default_bytes(str(DEFAULT_FILE))
source1, source2, source3 = io.BytesIO(source_bytes), io.BytesIO(source_bytes), io.BytesIO(source_bytes)

try:
    base_df = load_margin(source1)
    base_target, base_basis, base_rates = read_assumptions(source2)
    workbook_sheets = read_all_sheets(source3)
except Exception as e:
    st.error(f"Could not read the workbook: {e}")
    st.stop()

st.sidebar.markdown("### Model Controls")
target = st.sidebar.slider("Target net margin", 0.0, 0.80, float(base_target), 0.01, format="%.0f%%")
basis = st.sidebar.selectbox("Operating cost basis", ["Landing Price Excl GST", "Cost Price"], index=0 if base_basis != "Cost Price" else 1)

with st.sidebar.expander("Scenario: operating costs", expanded=False):
    scenario_rates = {}
    for ptype in sorted(base_df["Product Type"].dropna().unique()):
        st.markdown(f"**{ptype}**")
        defaults = base_rates.get(ptype, {"Transportation":0,"Warehouse":0,"Salary":0,"Additional Services":0})
        scenario_rates[ptype] = {
            "Transportation": st.number_input(f"Transport % · {ptype}", 0.0, 1.0, float(defaults.get("Transportation",0)), 0.01, format="%.2f", key=f"t_{ptype}"),
            "Warehouse": st.number_input(f"Warehouse % · {ptype}", 0.0, 1.0, float(defaults.get("Warehouse",0)), 0.01, format="%.2f", key=f"w_{ptype}"),
            "Salary": st.number_input(f"Salary % · {ptype}", 0.0, 1.0, float(defaults.get("Salary",0)), 0.01, format="%.2f", key=f"s_{ptype}"),
            "Additional Services": st.number_input(f"Services % · {ptype}", 0.0, 1.0, float(defaults.get("Additional Services",0)), 0.01, format="%.2f", key=f"a_{ptype}"),
        }

df = apply_model(base_df, scenario_rates, basis, target)

st.sidebar.markdown("### Filters")
ay_opts = sorted([x for x in df["Academic Year"].unique() if x != "Unknown"], reverse=True)
selected_ay = st.sidebar.multiselect("Academic Year", ay_opts, default=ay_opts)
prod_opts = sorted(df["Product Type"].unique())
selected_prod = st.sidebar.multiselect("Product Type", prod_opts, default=prod_opts)
grade_opts = sorted(df["Grades"].unique(), key=grade_num)
selected_grade = st.sidebar.multiselect("Grade", grade_opts, default=grade_opts)
subject_opts = sorted(df["Subjects"].unique())
selected_subject = st.sidebar.multiselect("Subject", subject_opts, default=subject_opts)

mask = df["Academic Year"].isin(selected_ay) & df["Product Type"].isin(selected_prod) & df["Grades"].isin(selected_grade) & df["Subjects"].isin(selected_subject)
f = df.loc[mask].copy()

st.markdown('''<div class="hero"><div class="eyebrow">FINANCE & PRODUCT ANALYTICS</div><h1>Academic Product Margin Intelligence</h1><p>One interactive view of landing value, cost structure, gross margin, operating costs and net profitability across academic years, products and grades.</p></div>''', unsafe_allow_html=True)

if f.empty:
    st.warning("No data matches the selected filters.")
    st.stop()

landing = f["Landing Price Excl GST (₹)"].sum()
cost = f["Cost Price (₹)"].sum()
gross = f["Gross Margin (₹)"].sum()
op = f["Total Operating Cost (₹)"].sum()
net = f["Net Margin (₹)"].sum()
gross_pct = gross / landing if landing else 0
net_pct = net / landing if landing else 0
below = int((f["Net Margin %"] < target).sum())

cols = st.columns(6)
with cols[0]: kpi("Landing Price Excl GST", money(landing), f"{len(f):,} line items", label_class="landing")
with cols[1]: kpi("Cost Price", money(cost), f"{(cost/landing*100 if landing else 0):.1f}% of landing", label_class="cost")
with cols[2]: kpi("Gross Margin", money(gross), f"{pct(gross_pct)} gross margin", "good" if gross_pct >= target else "warn")
with cols[3]: kpi("Operating Cost", money(op), f"{(op/landing*100 if landing else 0):.1f}% of landing")
with cols[4]: kpi("Net Margin", money(net), f"{pct(net_pct)} net margin", "good" if net_pct >= target else "bad")
with cols[5]: kpi("Below Target", f"{below:,}", f"Target: {pct(target)}", "good" if below == 0 else "warn")

st.write("")
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "Executive Overview", "Product & Grade", "Cost Structure", "Item Explorer",
    "Scenario Insights", "Workbook Explorer",
])

chart_layout = dict(
    margin=dict(l=10, r=10, t=45, b=10),
    height=370,
    legend_title_text="",
    font=dict(family="Manrope, sans-serif", color="#334155"),
    paper_bgcolor="rgba(255,255,255,0.72)",
    plot_bgcolor="rgba(255,255,255,0.35)",
    hoverlabel=dict(bgcolor="#0f172a", font=dict(color="white")),
)

with tab1:
    c1, c2 = st.columns([1.15, .85])
    by_prod = f.groupby("Product Type", as_index=False).agg(Landing=("Landing Price Excl GST (₹)","sum"), Net=("Net Margin (₹)","sum"), Gross=("Gross Margin (₹)","sum"), Operating=("Total Operating Cost (₹)","sum"))
    by_prod["Net Margin %"] = np.where(by_prod["Landing"] != 0, by_prod["Net"] / by_prod["Landing"], 0)
    with c1:
        st.markdown('<div class="section-title">Profitability by Product Type</div><div class="section-sub">Net margin value with margin-rate overlay</div>', unsafe_allow_html=True)
        fig = go.Figure()
        fig.add_bar(x=by_prod["Product Type"], y=by_prod["Net"], name="Net Margin (₹)")
        fig.add_trace(go.Scatter(x=by_prod["Product Type"], y=by_prod["Net Margin %"], name="Net Margin %", yaxis="y2", mode="lines+markers"))
        fig.update_layout(**chart_layout, yaxis=dict(title="Net Margin (₹)"), yaxis2=dict(title="Net Margin %", overlaying="y", side="right", tickformat=".0%"), hovermode="x unified")
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        st.markdown('<div class="section-title">Margin Health</div><div class="section-sub">Share of items meeting the selected target</div>', unsafe_allow_html=True)
        ok = len(f)-below
        donut = go.Figure(go.Pie(labels=["On / Above Target","Below Target"], values=[ok,below], hole=.70, textinfo="label+percent"))
        donut.update_layout(**chart_layout, showlegend=False, annotations=[dict(text=f"{ok/len(f):.0%}<br><span style='font-size:11px'>healthy</span>", x=.5,y=.5,showarrow=False,font_size=21)])
        st.plotly_chart(donut, use_container_width=True)

    by_year = f.groupby("Academic Year", as_index=False).agg(Landing=("Landing Price Excl GST (₹)","sum"), Product_Cost=("Cost Price (₹)","sum"), Operating_Cost=("Total Operating Cost (₹)","sum"), Net_Margin=("Net Margin (₹)","sum"))
    year_long = by_year.melt(id_vars="Academic Year", var_name="Metric", value_name="Value")
    st.markdown('<div class="section-title">Academic Year Value Bridge</div><div class="section-sub">Compare landing value against product, operating and net margin components</div>', unsafe_allow_html=True)
    fig = px.bar(year_long, x="Academic Year", y="Value", color="Metric", barmode="group")
    fig.update_layout(**chart_layout, yaxis_title="₹", xaxis_title="")
    st.plotly_chart(fig, use_container_width=True)

with tab2:
    by_grade = f.groupby(["Product Type","Grades"], as_index=False).agg(Items=("Item Name","size"), Landing=("Landing Price Excl GST (₹)","sum"), Net=("Net Margin (₹)","sum"))
    by_grade["Net Margin %"] = np.where(by_grade["Landing"] != 0, by_grade["Net"] / by_grade["Landing"], 0)
    by_grade["Grade Order"] = by_grade["Grades"].map(grade_num)
    by_grade = by_grade.sort_values(["Product Type","Grade Order"])
    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="section-title">Grade-wise Net Margin %</div><div class="section-sub">Identify grades with stronger or weaker unit economics</div>', unsafe_allow_html=True)
        fig = px.bar(by_grade, x="Grades", y="Net Margin %", color="Product Type", barmode="group", hover_data=["Items"])
        fig.add_hline(y=target, line_dash="dash", annotation_text="Target")
        fig.update_layout(**chart_layout, yaxis_tickformat=".0%", xaxis_title="", yaxis_title="Net Margin %")
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        st.markdown('<div class="section-title">Landing Value Mix</div><div class="section-sub">Relative value contribution by product and grade</div>', unsafe_allow_html=True)
        fig = px.treemap(by_grade, path=["Product Type","Grades"], values="Landing", color="Net Margin %", color_continuous_scale="RdYlGn", color_continuous_midpoint=target)
        fig.update_layout(**chart_layout)
        st.plotly_chart(fig, use_container_width=True)

    summary = by_grade[["Product Type","Grades","Items","Landing","Net","Net Margin %"]].copy()
    summary.columns = ["Product Type","Grade","Items","Landing Value (₹)","Net Margin (₹)","Net Margin %"]
    st.dataframe(summary.style.format({"Landing Value (₹)":"₹{:,.0f}","Net Margin (₹)":"₹{:,.0f}","Net Margin %":"{:.1%}"}), use_container_width=True, hide_index=True, height=430)

with tab3:
    cost_cols = ["Transportation Cost (₹)","Warehouse & Facility Cost (₹)","Employee Salary Cost (₹)","Additional Services Cost (₹)"]
    cost = f.groupby("Product Type")[cost_cols].sum().reset_index().melt(id_vars="Product Type", var_name="Cost Component", value_name="Amount")
    cost["Cost Component"] = cost["Cost Component"].str.replace(" Cost (₹)","",regex=False).str.replace("Employee Salary","Salary",regex=False).str.replace("Warehouse & Facility","Warehouse",regex=False)
    c1, c2 = st.columns([1.15,.85])
    with c1:
        st.markdown('<div class="section-title">Operating Cost Composition</div><div class="section-sub">Stacked cost burden by product type</div>', unsafe_allow_html=True)
        fig = px.bar(cost, x="Product Type", y="Amount", color="Cost Component", barmode="stack")
        fig.update_layout(**chart_layout, xaxis_title="", yaxis_title="₹")
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        st.markdown('<div class="section-title">Cost Mix</div><div class="section-sub">Overall operating-cost allocation</div>', unsafe_allow_html=True)
        pie = cost.groupby("Cost Component",as_index=False)["Amount"].sum()
        fig = px.pie(pie, names="Cost Component", values="Amount", hole=.48)
        fig.update_layout(**chart_layout)
        st.plotly_chart(fig, use_container_width=True)

    st.markdown('<div class="section-title">Scenario Rates Applied</div><div class="section-sub">Rates are editable from the left-side Model Controls</div>', unsafe_allow_html=True)
    rate_rows=[]
    for p,r in scenario_rates.items():
        rate_rows.append({"Product Type":p, **{k:v for k,v in r.items()}, "Total":sum(r.values())})
    rate_df=pd.DataFrame(rate_rows)
    st.dataframe(rate_df.style.format({c:"{:.1%}" for c in ["Transportation","Warehouse","Salary","Additional Services","Total"]}), use_container_width=True, hide_index=True)

with tab4:
    st.markdown('<div class="section-title">Item-level Margin Explorer</div><div class="section-sub">Search, sort and inspect the exact items driving profitability</div>', unsafe_allow_html=True)
    query = st.text_input("Search item / subject / SKU", placeholder="e.g. Robotics, English, 400001...")
    items = f.copy()
    if query:
        q=query.lower()
        search_cols=["Item Name","Subjects","Core Kit SKU","Product Name"]
        hit=pd.Series(False,index=items.index)
        for c in search_cols:
            if c in items: hit |= items[c].astype(str).str.lower().str.contains(q,na=False)
        items=items[hit]
    show_cols=["Academic Year","Product Type","Grades","Subjects","Item Name","Item Type","Core Kit SKU","Landing Price Excl GST (₹)","Cost Price (₹)","Gross Margin %","Total Operating Cost (₹)","Net Margin (₹)","Net Margin %","Net Margin Status"]
    show_cols=[c for c in show_cols if c in items.columns]
    sty=items[show_cols].sort_values("Net Margin %",ascending=True).style.format({
        "Landing Price Excl GST (₹)":"₹{:,.2f}","Cost Price (₹)":"₹{:,.2f}","Gross Margin %":"{:.1%}","Total Operating Cost (₹)":"₹{:,.2f}","Net Margin (₹)":"₹{:,.2f}","Net Margin %":"{:.1%}"})
    st.dataframe(sty, use_container_width=True, hide_index=True, height=565)
    csv=items[show_cols].to_csv(index=False).encode("utf-8")
    st.download_button("Download filtered item analysis (CSV)", csv, "margin_item_analysis.csv", "text/csv")

with tab5:
    st.markdown('<div class="section-title">Scenario & Management Insights</div><div class="section-sub">Automatically surface concentration, margin pressure and priority-review areas</div>', unsafe_allow_html=True)
    c1,c2,c3=st.columns(3)
    worst_prod=by_prod.sort_values("Net Margin %").iloc[0]
    best_prod=by_prod.sort_values("Net Margin %",ascending=False).iloc[0]
    high_cost = f.groupby("Product Type")["Total Operating Cost (₹)"].sum().sort_values(ascending=False)
    with c1:
        st.markdown(f'<div class="info-card"><b>Strongest product</b><br><span style="font-size:24px;font-weight:800">{best_prod["Product Type"]}</span><br><span class="good">{pct(best_prod["Net Margin %"])} net margin</span></div>',unsafe_allow_html=True)
    with c2:
        state="good" if worst_prod["Net Margin %"]>=target else "bad"
        st.markdown(f'<div class="info-card"><b>Priority margin review</b><br><span style="font-size:24px;font-weight:800">{worst_prod["Product Type"]}</span><br><span class="{state}">{pct(worst_prod["Net Margin %"])} net margin</span></div>',unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="info-card"><b>Largest operating-cost pool</b><br><span style="font-size:24px;font-weight:800">{high_cost.index[0]}</span><br><span>{money(high_cost.iloc[0])}</span></div>',unsafe_allow_html=True)

    risk = f[f["Net Margin %"] < target].groupby(["Product Type","Grades"],as_index=False).agg(Review_Items=("Item Name","size"), Net_Margin=("Net Margin (₹)","sum"), Landing=("Landing Price Excl GST (₹)","sum"))
    if not risk.empty:
        risk["Net Margin %"] = np.where(risk["Landing"]!=0,risk["Net_Margin"]/risk["Landing"],0)
        risk=risk.sort_values(["Review_Items","Net Margin %"],ascending=[False,True]).head(15)
        st.markdown("#### Top review clusters")
        st.dataframe(risk[["Product Type","Grades","Review_Items","Net Margin %","Net_Margin"]].style.format({"Net Margin %":"{:.1%}","Net_Margin":"₹{:,.0f}"}),use_container_width=True,hide_index=True)
    else:
        st.success("All selected items are on or above the current target margin.")

with tab6:
    st.markdown(
        '<div class="section-title">Workbook Explorer</div>'
        '<div class="section-sub">Browse every sheet and all columns from the uploaded Excel workbook</div>',
        unsafe_allow_html=True,
    )
    selected_sheet = st.selectbox("Sheet", list(workbook_sheets))
    sheet_df = workbook_sheets[selected_sheet]
    st.caption(f"{len(sheet_df):,} rows · {len(sheet_df.columns):,} columns")
    st.dataframe(sheet_df, use_container_width=True, hide_index=True, height=600)
    st.download_button(
        "Download selected sheet (CSV)",
        sheet_df.to_csv(index=False).encode("utf-8"),
        f"{selected_sheet}.csv",
        "text/csv",
    )

st.caption("Model logic: item-level landing value, gross margin and operating costs are recomputed from the workbook's raw fields and Assumptions sheet. Filters and scenario controls do not alter the source Excel file.")
